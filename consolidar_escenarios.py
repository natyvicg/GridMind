"""
consolidar_escenarios.py — Genera la tabla maestra de escenarios de validación.

NOTA: Este script ya fue ejecutado y sus salidas generadas. No necesita
volver a correrse salvo que se modifiquen los escenarios base. Requiere
el módulo metricas.py que fue parte del entorno de construcción original
y no está incluido en el repositorio final (las salidas ya generadas son
las que se usan en validación y documentación del TFG).

Salidas que genera:
    - escenarios_maestros.xlsx  (hoja resumen + 6 hojas detalle + hoja preguntas)
    - escenarios_maestros.md    (markdown espejo)
    - escenarios_maestros.pdf   (PDF listo para anexo del TFG)
"""
 
import os
import sys
import warnings
warnings.filterwarnings('ignore')
 
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
from escenarios import BUILDERS

# metricas.py fue parte del entorno de construcción original.
# Si necesitás re-ejecutar este script, reconstruí metricas.py con las
# funciones: extraer_metricas_globales, extraer_detalle_violaciones,
# elegir_elemento_n1, y las constantes LIMITE_V_ESTRICTO, LIMITE_V_LAXO.
try:
    from metricas import (extraer_metricas_globales, extraer_detalle_violaciones,
                          elegir_elemento_n1, LIMITE_V_ESTRICTO, LIMITE_V_LAXO)
except ImportError:
    print("AVISO: metricas.py no encontrado. Este script es histórico y ya fue ejecutado.")
    print("       Las salidas generadas están en la carpeta del proyecto.")
    sys.exit(0)
 
# ─── Batería de preguntas al agente ──────────────────────────────────────
PREGUNTAS_BATERIA = [
    ('Q1', 'Básica',
     '¿Converge el flujo de potencia? ¿Cuántas barras tiene la red?'),
    ('Q2', 'Diagnóstico tensión',
     '¿Hay violaciones de tensión? Listá las barras afectadas con su vm_pu.'),
    ('Q3', 'Diagnóstico térmico',
     '¿Hay líneas o trafos sobrecargados? Listá con su % de carga.'),
    ('Q4', 'Síntesis',
     'Resumí el estado operativo de la red en 3-4 líneas.'),
    ('Q5', 'Razonamiento',
     'Dadas las violaciones encontradas, ¿qué acción correctiva sugerirías?'),
    ('Q6', 'N-1',
     'Simulá la salida de la línea indicada y reportá el impacto.'),
]
 
# ─── Estilos Excel ────────────────────────────────────────────────────────
FILL_HEADER = PatternFill('solid', start_color='1F4E78')
FILL_SUBHEADER = PatternFill('solid', start_color='D9E1F2')
FILL_IEEE = PatternFill('solid', start_color='E2EFDA')
FILL_CR = PatternFill('solid', start_color='FCE4D6')
FILL_VIOL = PatternFill('solid', start_color='FFE699')
FONT_HEADER = Font(bold=True, color='FFFFFF', name='Arial', size=11)
FONT_SUBHEADER = Font(bold=True, name='Arial', size=10)
FONT_BODY = Font(name='Arial', size=10)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
 
THIN = Side(border_style='thin', color='999999')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
 
 
def fmt_cell(cell, bold=False, fill=None, center=False, font_size=10):
    cell.font = Font(bold=bold, name='Arial', size=font_size)
    if fill is not None:
        cell.fill = fill
    cell.alignment = CENTER if center else LEFT
    cell.border = BORDER
 
 
# ─── Hoja Resumen ─────────────────────────────────────────────────────────
def _escribir_hoja_resumen(wb, registros, elementos_n1):
    ws = wb.create_sheet('Resumen', 0)
    ws.sheet_view.showGridLines = False
    
    # Título
    ws['A1'] = 'TABLA MAESTRA DE ESCENARIOS DE VALIDACIÓN'
    ws['A1'].font = Font(bold=True, size=14, name='Arial')
    ws.merge_cells('A1:P1')
    ws['A1'].alignment = CENTER
    
    ws['A2'] = f'GridMind — Día 4 | Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = Font(italic=True, size=9, name='Arial', color='666666')
    ws.merge_cells('A2:P2')
    ws['A2'].alignment = CENTER
    
    # Encabezados de grupo (fila 4)
    grupos = [
        ('A4:F4', 'IDENTIFICACIÓN'),
        ('G4:G4', 'SETUP'),
        ('H4:L4', 'ESTADO GLOBAL'),
        ('M4:P4', 'VIOLACIONES'),
    ]
    for rng, txt in grupos:
        cell = ws[rng.split(':')[0]]
        cell.value = txt
        ws.merge_cells(rng)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = CENTER
        cell.border = BORDER
    
    # Encabezados de columna (fila 5)
    encabezados = [
        'ID', 'Red', 'Nombre', 'Modelo', 'Propósito', 'N-1 sugerido',
        'Condición inicial',
        'Conv.', 'Vmin (pu)', 'Vmax (pu)', 'L máx %', 'T máx %',
        '#V estricto\n(0.95-1.05)', '#V laxo\n(0.90-1.10)', '#L>100%', '#T>100%'
    ]
    for j, h in enumerate(encabezados, start=1):
        c = ws.cell(row=5, column=j, value=h)
        c.fill = FILL_SUBHEADER
        c.font = FONT_SUBHEADER
        c.alignment = CENTER
        c.border = BORDER
    
    # Filas de datos
    for i, (reg, n1) in enumerate(zip(registros, elementos_n1), start=6):
        fill = FILL_IEEE if 'IEEE' in reg['red'] else FILL_CR
        
        n1_txt = n1['descripcion'] if n1 else '—'
        
        valores = [
            reg['id'], reg['red'], reg['nombre'], reg.get('modelo', '—'),
            reg['proposito'], n1_txt,
            reg['condicion_inicial'],
            '✅' if reg['convergio'] else '❌',
            f"{reg['vmin_pu']:.4f}" if pd.notna(reg['vmin_pu']) else '—',
            f"{reg['vmax_pu']:.4f}" if pd.notna(reg['vmax_pu']) else '—',
            f"{reg['carga_max_linea_pct']:.1f}" if pd.notna(reg['carga_max_linea_pct']) else '—',
            f"{reg['carga_max_trafo_pct']:.1f}" if pd.notna(reg['carga_max_trafo_pct']) else '—',
            reg['n_viola_V_estricto'], reg['n_viola_V_laxo'],
            reg['n_lineas_over'], reg['n_trafos_over'],
        ]
        for j, v in enumerate(valores, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.fill = fill
            c.font = FONT_BODY
            c.alignment = CENTER if j not in (5, 6, 7) else LEFT
            c.border = BORDER
            # Resaltar violaciones
            if j in (13, 14, 15, 16) and isinstance(v, int) and v > 0:
                c.fill = FILL_VIOL
                c.font = Font(name='Arial', size=10, bold=True)
    
    # Anchos de columna
    anchos = [5, 12, 14, 13, 35, 40, 55, 7, 11, 11, 10, 10, 12, 12, 10, 10]
    for j, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[5].height = 35
    for i in range(6, 6 + len(registros)):
        ws.row_dimensions[i].height = 50
    
    # Nota al pie
    nota_fila = 6 + len(registros) + 2
    ws.cell(row=nota_fila, column=1, value='Notas:').font = Font(bold=True, size=9, name='Arial')
    notas = [
        ('Modelos IEEE 14: "Oficial" = CDF del AEP Test System 1962 sin modificaciones '
         '(referencia matemática). "Calibrado" = setpoints normalizados a 1.02 pu y '
         'max_i_ka recalibrado a 2× flujo base (condiciones operativas modernas).'),
        ('Modelo de Costa Rica: oficial del ICE. 524 barras declaradas, 513 energizadas '
         '(11 aisladas por despacho operativo — ver anexo "desvío del Día 4").'),
        (f'Límite estricto de tensión: {LIMITE_V_ESTRICTO[0]}–{LIMITE_V_ESTRICTO[1]} pu. '
         f'Límite laxo: {LIMITE_V_LAXO[0]}–{LIMITE_V_LAXO[1]} pu.'),
        'Carga máx. trafo 2W en IEEE 14 no aplica (la red solo tiene líneas).',
        'El detalle de cada escenario (barras/líneas violadas) está en las hojas E1-E7.',
    ]
    for k, n in enumerate(notas, start=1):
        c = ws.cell(row=nota_fila + k, column=1, value=f'  • {n}')
        c.font = Font(italic=True, size=9, name='Arial')
        ws.merge_cells(start_row=nota_fila + k, start_column=1,
                       end_row=nota_fila + k, end_column=16)
    
    ws.freeze_panes = 'A6'
 
 
# ─── Hojas de Detalle por escenario ───────────────────────────────────────
def _escribir_hoja_detalle(wb, reg, df_bus, df_line, df_trafo, df_trafo3, n1):
    sheet_name = f"{reg['id']}_{reg['nombre'].replace('Ó', 'O').replace('Í', 'I')[:20]}"
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    
    r = 1
    # Cabecera
    ws.cell(row=r, column=1, value=f"{reg['id']} — {reg['red']} — {reg['nombre']}").font = \
        Font(bold=True, size=14, name='Arial')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 2
    
    # Metadatos
    metadatos = [
        ('Red', reg['red']),
        ('Modelo', reg.get('modelo', '—')),
        ('Propósito de validación', reg['proposito']),
        ('Condición inicial', reg['condicion_inicial']),
        ('Barras totales', reg.get('n_barras_totales', '—')),
        ('Barras energizadas', reg.get('n_barras_energizadas', '—')),
        ('Barras con resultado PF', reg.get('n_barras_con_resultado', '—')),
        ('Convergencia', '✅ Sí' if reg['convergio'] else '❌ No'),
        ('Vmin (pu)', f"{reg['vmin_pu']:.4f}" if pd.notna(reg['vmin_pu']) else '—'),
        ('Vmax (pu)', f"{reg['vmax_pu']:.4f}" if pd.notna(reg['vmax_pu']) else '—'),
        ('Carga máx. línea (%)', f"{reg['carga_max_linea_pct']:.2f}"),
        ('Elemento N-1 sugerido', n1['descripcion'] if n1 else '—'),
    ]
    for k, v in metadatos:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, name='Arial', size=10)
        ws.cell(row=r, column=2, value=str(v)).font = Font(name='Arial', size=10)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        r += 1
    r += 1
    
    # Sección barras violadas
    ws.cell(row=r, column=1, value='BARRAS EN VIOLACIÓN DE TENSIÓN').font = \
        Font(bold=True, size=12, name='Arial', color='FFFFFF')
    ws.cell(row=r, column=1).fill = FILL_HEADER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
    
    if df_bus.empty:
        ws.cell(row=r, column=1, value='— Sin violaciones de tensión (estricto ni laxo) —').font = \
            Font(italic=True, name='Arial', size=10)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        r += 2
    else:
        cols = list(df_bus.columns)
        for j, col in enumerate(cols, start=1):
            c = ws.cell(row=r, column=j, value=col)
            c.fill = FILL_SUBHEADER
            c.font = FONT_SUBHEADER
            c.border = BORDER
        r += 1
        for _, row in df_bus.iterrows():
            for j, col in enumerate(cols, start=1):
                v = row[col]
                c = ws.cell(row=r, column=j, value=v)
                c.font = FONT_BODY
                c.border = BORDER
                if col == 'severidad' and v == 'estricto':
                    c.fill = FILL_VIOL
            r += 1
        r += 1
    
    # Sección líneas sobrecargadas
    ws.cell(row=r, column=1, value='LÍNEAS SOBRECARGADAS (>100%)').font = \
        Font(bold=True, size=12, name='Arial', color='FFFFFF')
    ws.cell(row=r, column=1).fill = FILL_HEADER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
    
    if df_line.empty:
        ws.cell(row=r, column=1, value='— Sin líneas sobrecargadas —').font = \
            Font(italic=True, name='Arial', size=10)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        r += 2
    else:
        cols = list(df_line.columns)
        for j, col in enumerate(cols, start=1):
            c = ws.cell(row=r, column=j, value=col)
            c.fill = FILL_SUBHEADER
            c.font = FONT_SUBHEADER
            c.border = BORDER
        r += 1
        for _, row in df_line.iterrows():
            for j, col in enumerate(cols, start=1):
                c = ws.cell(row=r, column=j, value=row[col])
                c.font = FONT_BODY
                c.border = BORDER
                if col == 'loading_pct':
                    c.fill = FILL_VIOL
            r += 1
        r += 1
    
    # Sección trafos sobrecargados
    ws.cell(row=r, column=1, value='TRANSFORMADORES SOBRECARGADOS (>100%)').font = \
        Font(bold=True, size=12, name='Arial', color='FFFFFF')
    ws.cell(row=r, column=1).fill = FILL_HEADER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
    
    if df_trafo.empty and df_trafo3.empty:
        ws.cell(row=r, column=1, value='— Sin trafos sobrecargados —').font = \
            Font(italic=True, name='Arial', size=10)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        r += 2
    else:
        if not df_trafo.empty:
            ws.cell(row=r, column=1, value='Trafos 2W').font = Font(bold=True, name='Arial')
            r += 1
            cols = list(df_trafo.columns)
            for j, col in enumerate(cols, start=1):
                c = ws.cell(row=r, column=j, value=col)
                c.fill = FILL_SUBHEADER; c.font = FONT_SUBHEADER; c.border = BORDER
            r += 1
            for _, row in df_trafo.iterrows():
                for j, col in enumerate(cols, start=1):
                    c = ws.cell(row=r, column=j, value=row[col])
                    c.font = FONT_BODY; c.border = BORDER
                r += 1
            r += 1
        if not df_trafo3.empty:
            ws.cell(row=r, column=1, value='Trafos 3W').font = Font(bold=True, name='Arial')
            r += 1
            cols = list(df_trafo3.columns)
            for j, col in enumerate(cols, start=1):
                c = ws.cell(row=r, column=j, value=col)
                c.fill = FILL_SUBHEADER; c.font = FONT_SUBHEADER; c.border = BORDER
            r += 1
            for _, row in df_trafo3.iterrows():
                for j, col in enumerate(cols, start=1):
                    c = ws.cell(row=r, column=j, value=row[col])
                    c.font = FONT_BODY; c.border = BORDER
                r += 1
            r += 1
    
    # Anchos
    for j in range(1, 8):
        ws.column_dimensions[get_column_letter(j)].width = 18
    ws.column_dimensions['A'].width = 22
 
 
# ─── Hoja Preguntas ───────────────────────────────────────────────────────
def _escribir_hoja_preguntas(wb):
    ws = wb.create_sheet('Preguntas')
    ws.sheet_view.showGridLines = False
    
    ws['A1'] = 'BATERÍA DE VALIDACIÓN — PREGUNTAS AL AGENTE'
    ws['A1'].font = Font(bold=True, size=14, name='Arial')
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = CENTER
    
    ws['A2'] = ('Batería única aplicada a los 6 escenarios. '
                'Las mismas preguntas se ejecutan en modo single-shot y en modo conversación (multi-turn).')
    ws['A2'].font = Font(italic=True, size=10, name='Arial', color='666666')
    ws.merge_cells('A2:D2')
    ws['A2'].alignment = LEFT
    
    # Encabezado
    headers = ['ID', 'Tipo', 'Pregunta', 'Criterio de evaluación']
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.fill = FILL_HEADER; c.font = FONT_HEADER; c.alignment = CENTER; c.border = BORDER
    
    criterios = {
        'Q1': 'Determinístico: convergencia (sí/no) + #barras exacto',
        'Q2': 'Determinístico: lista exacta de barras violadas con vm_pu (tolerancia 0.001 pu)',
        'Q3': 'Determinístico: lista exacta de líneas/trafos con % (tolerancia 0.1%)',
        'Q4': 'Cualitativo: coherencia con las respuestas a Q1-Q3',
        'Q5': 'Semi-determinístico en IEEE 14 (respuesta esperada acotada). Cualitativo en CR',
        'Q6': 'Determinístico: recomputar PF con línea fuera y comparar resultados',
    }
    
    for i, (qid, tipo, pregunta) in enumerate(PREGUNTAS_BATERIA, start=5):
        ws.cell(row=i, column=1, value=qid).font = Font(bold=True, name='Arial', size=10)
        ws.cell(row=i, column=2, value=tipo).font = FONT_BODY
        ws.cell(row=i, column=3, value=pregunta).font = FONT_BODY
        ws.cell(row=i, column=4, value=criterios[qid]).font = FONT_BODY
        for j in range(1, 5):
            ws.cell(row=i, column=j).border = BORDER
            ws.cell(row=i, column=j).alignment = LEFT
    
    # Anchos
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 70
    ws.column_dimensions['D'].width = 60
    for i in range(5, 5 + len(PREGUNTAS_BATERIA)):
        ws.row_dimensions[i].height = 40
 
 
# ─── XLSX ────────────────────────────────────────────────────────────────
def generar_xlsx(registros, detalles, elementos_n1, path='escenarios_maestros.xlsx'):
    wb = Workbook()
    wb.remove(wb.active)
    
    _escribir_hoja_resumen(wb, registros, elementos_n1)
    for reg, (df_bus, df_line, df_trafo, df_trafo3), n1 in zip(registros, detalles, elementos_n1):
        _escribir_hoja_detalle(wb, reg, df_bus, df_line, df_trafo, df_trafo3, n1)
    _escribir_hoja_preguntas(wb)
    
    wb.save(path)
    return path
 
 
# ─── Markdown ─────────────────────────────────────────────────────────────
def generar_md(registros, detalles, elementos_n1, path='escenarios_maestros.md'):
    lines = []
    lines.append('# Tabla Maestra de Escenarios de Validación — GridMind')
    lines.append('')
    lines.append(f'*Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}*')
    lines.append('')
    lines.append('## 1. Contexto')
    lines.append('')
    lines.append('Este documento consolida los siete escenarios sobre los que se valida GridMind. '
                 'Cuatro son sobre la red IEEE 14 barras y tres corresponden a la red real de '
                 'transmisión de Costa Rica en los despachos Min, Med y Max de marzo 2023.')
    lines.append('')
    lines.append('**Modelos de la red IEEE 14.** Se distinguen dos grupos:')
    lines.append('')
    lines.append('- **Oficial** (E1 ORIGINAL): modelo CDF del AEP Test System (1962) sin '
                 'modificaciones. Setpoints 1.06–1.09 pu, max_i_ka = 42.33 kA. Es referencia '
                 'matemática para validación de solvers, no cumple criterios operativos modernos.')
    lines.append('- **Calibrado** (E2 BASE, E3 SUBTENSIÓN, E4 SOBRECARGA): sobre la red IEEE 14 se '
                 'normalizan los setpoints de gens a 1.02 pu y se recalibra max_i_ka a 2× el flujo '
                 'base. Representa condiciones operativas modernas (ANSI C84.1 / IEEE Std).')
    lines.append('')
    lines.append(f'**Límites de tensión:** estricto {LIMITE_V_ESTRICTO[0]}–{LIMITE_V_ESTRICTO[1]} pu, '
                 f'laxo {LIMITE_V_LAXO[0]}–{LIMITE_V_LAXO[1]} pu. '
                 'Se reportan ambos conteos para el modelo de CR, que tiene tensiones fuera del '
                 'rango estricto por diseño del despacho operativo.')
    lines.append('')
    lines.append('**Red de Costa Rica:** 524 barras declaradas, 513 energizadas. '
                 'Las 11 restantes permanecen aisladas por despacho operativo (ver anexo '
                 '"Desvío del Día 4 — Diagnóstico de 11 barras").')
    lines.append('')
    
    lines.append('## 2. Resumen de escenarios')
    lines.append('')
    lines.append('| ID | Red | Nombre | Modelo | Propósito | Conv. | Vmin | Vmax | L máx % | #V estricto | #V laxo | #L>100% |')
    lines.append('|---|---|---|---|---|---|---|---|---|---|---|---|')
    for reg in registros:
        lines.append(
            f"| {reg['id']} | {reg['red']} | {reg['nombre']} | "
            f"{reg.get('modelo', '—')} | "
            f"{reg['proposito'][:45]}{'…' if len(reg['proposito'])>45 else ''} | "
            f"{'✅' if reg['convergio'] else '❌'} | "
            f"{reg['vmin_pu']:.4f} | {reg['vmax_pu']:.4f} | "
            f"{reg['carga_max_linea_pct']:.1f} | "
            f"{reg['n_viola_V_estricto']} | {reg['n_viola_V_laxo']} | "
            f"{reg['n_lineas_over']} |"
        )
    lines.append('')
    
    lines.append('## 3. Condiciones iniciales')
    lines.append('')
    for reg, n1 in zip(registros, elementos_n1):
        lines.append(f"### {reg['id']} — {reg['red']} — {reg['nombre']}")
        lines.append('')
        lines.append(f"- **Modelo:** {reg.get('modelo', '—')}")
        lines.append(f"- **Propósito de validación:** {reg['proposito']}")
        lines.append(f"- **Condición inicial:** {reg['condicion_inicial']}")
        lines.append(f"- **Barras:** {reg.get('n_barras_totales', '—')} declaradas, "
                     f"{reg.get('n_barras_energizadas', '—')} energizadas")
        lines.append(f"- **Elemento N-1 sugerido (para Q6):** {n1['descripcion'] if n1 else '—'}")
        lines.append('')
    
    lines.append('## 4. Detalle de violaciones por escenario')
    lines.append('')
    for reg, (df_bus, df_line, df_trafo, df_trafo3) in zip(registros, detalles):
        lines.append(f"### {reg['id']} — {reg['nombre']}")
        lines.append('')
        
        # Barras
        if df_bus.empty:
            lines.append('**Tensiones:** sin violaciones.')
        else:
            lines.append(f"**Tensiones:** {len(df_bus)} barras en violación.")
            lines.append('')
            # Mostrar solo primeras 10 en el md
            n_mostrar = min(10, len(df_bus))
            lines.append(f'| bus_idx | name | vn_kv | vm_pu | categoría | severidad |')
            lines.append(f'|---|---|---|---|---|---|')
            for _, row in df_bus.head(n_mostrar).iterrows():
                lines.append(f"| {row['bus_idx']} | {row['name']} | {row['vn_kv']:.2f} | "
                             f"{row['vm_pu']:.4f} | {row['categoria']} | {row['severidad']} |")
            if len(df_bus) > n_mostrar:
                lines.append('')
                lines.append(f'*({len(df_bus) - n_mostrar} filas adicionales en la hoja detalle del xlsx.)*')
        lines.append('')
        
        # Líneas
        if df_line.empty:
            lines.append('**Líneas sobrecargadas:** ninguna.')
        else:
            lines.append(f"**Líneas sobrecargadas:** {len(df_line)}.")
            lines.append('')
            lines.append('| line_idx | from_bus | to_bus | long_km | loading % |')
            lines.append('|---|---|---|---|---|')
            for _, row in df_line.iterrows():
                lines.append(f"| {int(row['line_idx'])} | {int(row['from_bus'])} | "
                             f"{int(row['to_bus'])} | "
                             f"{row['longitud_km']:.2f} | {row['loading_pct']:.2f} |")
        lines.append('')
        
        # Trafos
        n_trafos_total = len(df_trafo) + len(df_trafo3)
        if n_trafos_total == 0:
            if reg['red'] == 'Costa Rica':
                lines.append('**Trafos sobrecargados:** ninguno.')
            lines.append('')
        else:
            lines.append(f"**Trafos sobrecargados:** {n_trafos_total}.")
            if not df_trafo.empty:
                lines.append('')
                lines.append('Trafos 2W:')
                lines.append('')
                lines.append('| trafo_idx | hv_bus | lv_bus | sn_mva | loading % |')
                lines.append('|---|---|---|---|---|')
                for _, row in df_trafo.iterrows():
                    lines.append(f"| {int(row['trafo_idx'])} | {int(row['hv_bus'])} | "
                                 f"{int(row['lv_bus'])} | "
                                 f"{row['sn_mva']:.1f} | {row['loading_pct']:.2f} |")
            if not df_trafo3.empty:
                lines.append('')
                lines.append('Trafos 3W:')
                lines.append('')
                lines.append('| trafo3w_idx | hv_bus | mv_bus | lv_bus | loading % |')
                lines.append('|---|---|---|---|---|')
                for _, row in df_trafo3.iterrows():
                    lines.append(f"| {int(row['trafo3w_idx'])} | {int(row['hv_bus'])} | "
                                 f"{int(row['mv_bus'])} | {int(row['lv_bus'])} | "
                                 f"{row['loading_pct']:.2f} |")
            lines.append('')
    
    lines.append('## 5. Batería de preguntas al agente')
    lines.append('')
    lines.append('Las mismas seis preguntas se aplican a los siete escenarios. '
                 'Cada escenario se ejecuta en dos modos: **single-shot** '
                 '(cada pregunta con contexto limpio) y **multi-turn** '
                 '(las seis preguntas en una misma conversación, probando persistencia de estado).')
    lines.append('')
    lines.append('| ID | Tipo | Pregunta | Criterio de evaluación |')
    lines.append('|---|---|---|---|')
    criterios = {
        'Q1': 'Determinístico: convergencia + #barras exacto',
        'Q2': 'Determinístico: lista exacta de barras (tol. 0.001 pu)',
        'Q3': 'Determinístico: lista exacta de líneas/trafos (tol. 0.1%)',
        'Q4': 'Cualitativo: coherencia con Q1-Q3',
        'Q5': 'Semi-determinístico (IEEE 14) / cualitativo (CR)',
        'Q6': 'Determinístico: recomputar PF y comparar',
    }
    for qid, tipo, preg in PREGUNTAS_BATERIA:
        lines.append(f'| {qid} | {tipo} | {preg} | {criterios[qid]} |')
    lines.append('')
    
    lines.append('## 6. Protocolo de validación (Día 7)')
    lines.append('')
    lines.append('1. Para cada escenario E1–E7, construir la red aplicando la condición inicial correspondiente.')
    lines.append('2. **Modo single-shot:** formular cada una de las 6 preguntas en una conversación independiente. Registrar tool calls, respuesta final y tiempo.')
    lines.append('3. **Modo multi-turn:** formular las 6 preguntas en secuencia dentro de una misma conversación. Validar que el estado de la red se preserva entre turnos.')
    lines.append('4. Comparar cada respuesta contra el ground truth de la sección 4.')
    lines.append('5. Clasificar el resultado por pregunta como *correcto*, *parcialmente correcto* o *incorrecto*, según el criterio de evaluación aplicable.')
    lines.append('')
    
    content = '\n'.join(lines)
    Path(path).write_text(content, encoding='utf-8')
    return path
 
 
# ─── PDF ─────────────────────────────────────────────────────────────────
def generar_pdf(md_path, pdf_path='escenarios_maestros.pdf'):
    """Convierte el markdown a PDF vía HTML (pandoc → HTML → wkhtmltopdf).
    Si las herramientas no están instaladas, avisa y devuelve None.
    """
    import subprocess
    import shutil
    
    # Verificar que las herramientas estén disponibles
    pandoc_ok = shutil.which('pandoc') is not None
    wkhtml_ok = shutil.which('wkhtmltopdf') is not None
    
    if not pandoc_ok or not wkhtml_ok:
        faltantes = []
        if not pandoc_ok:
            faltantes.append('pandoc')
        if not wkhtml_ok:
            faltantes.append('wkhtmltopdf')
        print(f"  ⚠ PDF no generado: faltan herramientas ({', '.join(faltantes)}).")
        print(f"    Instrucciones de instalación en Windows:")
        print(f"    - pandoc:     https://pandoc.org/installing.html")
        print(f"    - wkhtmltopdf: https://wkhtmltopdf.org/downloads.html")
        print(f"    Alternativa sin instalar nada: abrí el xlsx en Excel y exportá como PDF (Archivo → Exportar).")
        return None
    
    css = """
    body { font-family: 'DejaVu Sans', Arial, sans-serif; font-size: 10pt;
           max-width: 900px; margin: 2em auto; color: #222; line-height: 1.5; }
    h1 { color: #1f4e78; border-bottom: 2px solid #1f4e78; padding-bottom: 6px; }
    h2 { color: #1f4e78; margin-top: 1.5em; }
    h3 { color: #2e75b6; margin-top: 1.2em; }
    table { border-collapse: collapse; margin: 1em 0; font-size: 9pt; width: 100%; }
    th { background: #1f4e78; color: white; padding: 6px 8px; text-align: left; }
    td { border: 1px solid #bbb; padding: 5px 7px; vertical-align: top; }
    tr:nth-child(even) { background: #f5f5f5; }
    code { background: #eef; padding: 1px 4px; border-radius: 3px; font-size: 9pt; }
    em { color: #666; }
    """
    
    html_path = str(pdf_path).replace('.pdf', '.html')
    r1 = subprocess.run(
        ['pandoc', str(md_path), '-o', html_path,
         '-s', '--toc', '--toc-depth=2',
         '--metadata', 'title=Tabla Maestra de Escenarios — GridMind'],
        capture_output=True, text=True
    )
    if r1.returncode != 0:
        print(f"  ⚠ Warning pandoc md→html: {r1.stderr[:200]}")
        return None
    
    from pathlib import Path as _P
    html_content = _P(html_path).read_text(encoding='utf-8')
    html_content = html_content.replace('</head>', f'<style>{css}</style></head>')
    _P(html_path).write_text(html_content, encoding='utf-8')
    
    r2 = subprocess.run(
        ['wkhtmltopdf', '--quiet',
         '--margin-top', '15mm', '--margin-bottom', '15mm',
         '--margin-left', '15mm', '--margin-right', '15mm',
         '--encoding', 'UTF-8',
         '--enable-local-file-access',
         html_path, str(pdf_path)],
        capture_output=True, text=True
    )
    if r2.returncode != 0:
        print(f"  ⚠ Warning wkhtmltopdf: {r2.stderr[:200]}")
        return None
    return pdf_path
 
 
# ─── Main ─────────────────────────────────────────────────────────────────
def main(output_dir='.'):
    print(f'Construyendo los {len(BUILDERS)} escenarios y extrayendo métricas...')
    
    registros = []
    detalles = []
    elementos_n1 = []
    
    for i, builder in enumerate(BUILDERS, start=1):
        print(f'  [{i}/{len(BUILDERS)}] {builder.__name__} ...', end=' ', flush=True)
        net, meta = builder()
        metricas = extraer_metricas_globales(net, meta)
        detalle = extraer_detalle_violaciones(net)
        n1 = elegir_elemento_n1(net)
        registros.append(metricas)
        detalles.append(detalle)
        elementos_n1.append(n1)
        print(f'OK ({metricas["n_barras_con_resultado"]} barras con resultado, '
              f'Vmin={metricas["vmin_pu"]:.4f})')
    
    out = Path(output_dir)
    out.mkdir(exist_ok=True, parents=True)
    
    print('\nGenerando xlsx...')
    xlsx_path = generar_xlsx(registros, detalles, elementos_n1,
                              path=out / 'escenarios_maestros.xlsx')
    print(f'  ✓ {xlsx_path}')
    
    print('Generando markdown...')
    md_path = generar_md(registros, detalles, elementos_n1,
                         path=out / 'escenarios_maestros.md')
    print(f'  ✓ {md_path}')
    
    # print('Generando PDF...')
    # pdf_path = generar_pdf(md_path, pdf_path=out / 'escenarios_maestros.pdf')
    # if pdf_path:
    #     print(f'  ✓ {pdf_path}')
    
    print('\n✅ Tabla maestra consolidada.')
    return registros, detalles, elementos_n1
 
 
if __name__ == '__main__':
    main()